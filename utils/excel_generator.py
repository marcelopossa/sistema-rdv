from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def gerar_excel_rdv(viagem):
    """
    Gera planilha Excel no formato do seu RDV atual
    """
    try:
        print(f"📊 Gerando Excel para viagem ID: {viagem.id}")
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "RDV"
        
        # Configurar larguras das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Estilos
        titulo_font = Font(size=14, bold=True)
        header_font = Font(size=11, bold=True)
        normal_font = Font(size=10)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Linha 1: Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "RELATÓRIO DE DESPESAS DE VIAGEM"
        ws['A1'].font = titulo_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Linha 3: Cabeçalho do relatório
        row = 3
        ws[f'A{row}'] = "Cliente:"
        ws[f'B{row}'] = viagem.cliente.nome
        ws[f'D{row}'] = "Código:"
        ws[f'E{row}'] = f"2025.{viagem.data_viagem.strftime('%m.%d')} - {viagem.projeto or 'MP'}"
        
        row += 1
        ws[f'A{row}'] = "Projeto:"
        ws[f'B{row}'] = viagem.projeto or '-'
        ws[f'D{row}'] = "Período:"
        ws[f'E{row}'] = viagem.data_viagem.strftime('%d/%m/%Y')
        
        row += 2
        ws[f'A{row}'] = "Participantes da Despesa"
        ws[f'B{row}'] = viagem.participantes
        
        row += 1
        ws[f'A{row}'] = "Beneficiário do RDV"
        ws[f'B{row}'] = viagem.beneficiario
        
        # Linha da tabela de despesas
        row += 3
        
        # Cabeçalho da tabela
        headers = ['', 'Hospedagem', 'Refeições', 'Passagens', 'Taxi / App', 'Pedágio', 'Qtd. KM rodado', 'Total do KM rodado', 'Outros', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Dados da viagem
        row += 1
        ws.cell(row=row, column=1, value=viagem.data_viagem.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=f"R$ {viagem.valor_hospedagem:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=3, value=f"R$ {viagem.valor_alimentacao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=4, value="R$ 0,00")  # Passagens
        ws.cell(row=row, column=5, value="R$ 0,00")  # Taxi
        ws.cell(row=row, column=6, value=f"R$ {viagem.valor_pedagio:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=7, value=int(viagem.km_rodado))
        ws.cell(row=row, column=8, value=f"R$ {viagem.total_km:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=9, value="R$ 0,00")  # Outros
        ws.cell(row=row, column=10, value=f"R$ {viagem.total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        # Aplicar bordas e alinhamento
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.font = normal_font
        
        # Linha de soma
        row += 1
        ws.cell(row=row, column=1, value="Soma")
        ws.cell(row=row, column=2, value=f"R$ {viagem.valor_hospedagem:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=3, value=f"R$ {viagem.valor_alimentacao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=4, value="R$ 0,00")
        ws.cell(row=row, column=5, value="R$ 0,00")
        ws.cell(row=row, column=6, value=f"R$ {viagem.valor_pedagio:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=7, value=int(viagem.km_rodado))
        ws.cell(row=row, column=8, value=f"R$ {viagem.total_km:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=9, value="R$ 0,00")
        ws.cell(row=row, column=10, value=f"R$ {viagem.total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        # Estilo para linha de soma
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.font = header_font
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Linha de adiantamento
        row += 1
        ws.cell(row=row, column=1, value="Adiantamento*")
        ws.cell(row=row, column=10, value="R$ 0,00")
        
        # Linha total final
        row += 1
        ws.cell(row=row, column=1, value="Total")
        ws.cell(row=row, column=2, value=f"R$ {viagem.valor_hospedagem:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=3, value=f"R$ {viagem.valor_alimentacao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=4, value="R$ 0,00")
        ws.cell(row=row, column=5, value="R$ 0,00")
        ws.cell(row=row, column=6, value=f"R$ {viagem.valor_pedagio:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=7, value=int(viagem.km_rodado))
        ws.cell(row=row, column=8, value=f"R$ {viagem.total_km:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        ws.cell(row=row, column=9, value="R$ 0,00")
        ws.cell(row=row, column=10, value=f"R$ {viagem.total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        # Estilo para total final
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Informações finais
        row += 3
        ws[f'A{row}'] = "Equipe:"
        ws[f'D{row}'] = "Observações:"
        
        row += 2
        ws[f'A{row}'] = "Aprovação:"
        
        # Salvar arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"RDV_{viagem.cliente.nome}_{viagem.data_viagem.strftime('%Y%m%d')}_{timestamp}.xlsx"
        filepath = os.path.join('static/uploads', filename)
        
        wb.save(filepath)
        print(f"✅ Excel gerado: {filepath}")
        
        return filepath, filename
        
    except Exception as e:
        print(f"❌ Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()
        return None, None
